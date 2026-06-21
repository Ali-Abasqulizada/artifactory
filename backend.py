from collections import defaultdict
from flask import Flask, request, send_file, render_template, jsonify
import os, sqlite3, hashlib, json, difflib, html
import re
import hmac

import mimetypes
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

from library import startup
from library.config import config

from datetime import date, datetime, time
import openpyxl
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config.from_object(config.DevelopmentConfig)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app.config["STORAGE_FOLDER"] = os.path.join(BASE_DIR, app.config["STORAGE_FOLDER"])
app.config["DATABASE"] = os.path.join(BASE_DIR, app.config["DATABASE"])

startup.check_db_exists_or_fail(app.config["DATABASE"])
startup.create_folder(app.config["STORAGE_FOLDER"])

TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".py",
    ".js",
    ".ts",
    ".html",
    ".css",
    ".json",
    ".xml",
    ".yml",
    ".yaml",
    ".csv",
    ".sql",
    ".ini",
    ".cfg",
    ".conf",
    ".log",
    ".java",
    ".c",
    ".cpp",
    ".h",
    ".hpp",
    ".cs",
    ".php",
    ".rb",
    ".go",
    ".rs",
    ".sh",
    ".bat",
    ".ps1",
}
SQLITE_EXTENSIONS = {".db", ".sqlite", ".sqlite3"}
PDF_EXTENSIONS = {".pdf"}
DOCX_EXTENSIONS = {".docx"}
DOC_EXTENSIONS = {".doc"}
ZIP_EXTENSIONS = {".zip"}

XLSX_EXTENSIONS = {".xlsx"}
XLS_EXTENSIONS = {".xls"}

WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
WORD_TAG = f"{{{WORD_NS}}}"

DOCX_NAMESPACE = {"w": WORD_NS}

ABBREVIATIONS = {
    "mr",
    "mrs",
    "ms",
    "dr",
    "prof",
    "sr",
    "jr",
    "st",
    "vs",
    "etc",
    "e.g",
    "i.e",
    "fig",
    "no",
    "vol",
    "pp",
    "p",
}


def get_artifact_file_path(name, commit_hash):
    """
    Finds the real uploaded artifact path from the database.
    """
    conn, err = startup.connect_db(app.config["DATABASE"])
    if not conn:
        return None, str(err)

    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT artifact_storage.path
            FROM artifacts
            JOIN artifact_storage ON artifact_storage.artifact_id = artifacts.id
            WHERE artifacts.name = ? AND artifacts.commit_hash = ?
            ORDER BY artifacts.id DESC
            LIMIT 1
            """,
            (name, commit_hash),
        )

        row = cursor.fetchone()
        cursor.close()

        if not row:
            return (
                None,
                f"Commit hash '{commit_hash}' does not exist for artifact '{name}'",
            )

        return row[0], None

    except Exception as err:
        return None, f"Failed to find artifact path: {err}"

    finally:
        conn.close()


def normalize_excel_value(value):
    """
    Converts Excel cell values into stable text for comparison.

    It ignores Excel styling and focuses only on real cell content:
    - text
    - numbers
    - dates
    - times
    - formulas
    """
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, time):
        return value.strftime("%H:%M:%S")

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)

    text = str(value)
    text = text.replace("\xa0", " ")
    text = text.replace("\u200b", "")
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def get_xlsx_cell_map(file_path):
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=False)

    try:
        workbook_data = {}

        for sheet in workbook.worksheets:
            sheet_data = {}

            for row in sheet.iter_rows():
                for cell in row:
                    value_text = normalize_excel_value(cell.value)

                    if not value_text:
                        continue

                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        value_text = f"FORMULA: {value_text}"

                    sheet_data[(cell.row, cell.column)] = value_text

            workbook_data[sheet.title] = sheet_data

        return workbook_data, ""

    except Exception as err:
        return None, f"Failed to read XLSX file '{file_path}': {err}"

    finally:
        workbook.close()


def get_excel_cell_status(old_value, new_value):
    if old_value == new_value:
        return "same"

    if old_value and not new_value:
        return "removed"

    if not old_value and new_value:
        return "added"

    return "changed"


def calculate_xlsx_diff_table(file_path_a, file_path_b):
    workbook_a, err = get_xlsx_cell_map(file_path_a)
    if err:
        return None, err

    workbook_b, err = get_xlsx_cell_map(file_path_b)
    if err:
        return None, err

    sheet_names = sorted(set(workbook_a.keys()) | set(workbook_b.keys()))

    sheets = []
    total_changed_cells = 0

    for sheet_name in sheet_names:
        sheet_a = workbook_a.get(sheet_name, {})
        sheet_b = workbook_b.get(sheet_name, {})

        all_cells = set(sheet_a.keys()) | set(sheet_b.keys())

        changed_cells = [
            cell_coordinate
            for cell_coordinate in all_cells
            if sheet_a.get(cell_coordinate, "") != sheet_b.get(cell_coordinate, "")
        ]

        if not changed_cells:
            continue

        total_changed_cells += len(changed_cells)

        changed_rows = [row for row, column in changed_cells]
        changed_columns = [column for row, column in changed_cells]

        min_row = max(min(changed_rows) - 1, 1)
        max_row = max(changed_rows) + 1

        min_column = max(min(changed_columns) - 1, 1)
        max_column = max(changed_columns) + 1

        columns = [
            get_column_letter(column_number)
            for column_number in range(min_column, max_column + 1)
        ]

        rows = []

        for row_number in range(min_row, max_row + 1):
            row_cells = []

            for column_number in range(min_column, max_column + 1):
                coordinate = (row_number, column_number)

                old_value = sheet_a.get(coordinate, "")
                new_value = sheet_b.get(coordinate, "")

                row_cells.append(
                    {
                        "old_value": old_value,
                        "new_value": new_value,
                        "status": get_excel_cell_status(old_value, new_value),
                    }
                )

            rows.append({"row_number": row_number, "cells": row_cells})

        sheets.append(
            {
                "name": sheet_name,
                "changed_count": len(changed_cells),
                "columns": columns,
                "rows": rows,
            }
        )

    rendered_html = render_template(
        "excel_diff_table.html", sheets=sheets, total_changed_cells=total_changed_cells
    )

    return rendered_html, ""


def get_xlsx_lines(file_path):
    """
    Extracts simple content-based comparison lines from an .xlsx file.

    It compares:
    - sheet names
    - cell addresses
    - cell text
    - numbers
    - dates
    - formulas

    It ignores:
    - font
    - color
    - bold/italic
    - borders
    - background color
    - column width
    - row height
    - Excel styling
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=False)

        lines = []

        for sheet in workbook.worksheets:
            lines.append(f"--- Sheet: {sheet.title} ---")

            sheet_has_content = False

            for row in sheet.iter_rows():
                for cell in row:
                    value_text = normalize_excel_value(cell.value)

                    if not value_text:
                        continue

                    sheet_has_content = True

                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        line = (
                            f"{sheet.title}!{cell.coordinate} = FORMULA: {value_text}"
                        )
                    else:
                        line = f"{sheet.title}!{cell.coordinate} = {value_text}"

                    lines.append(line)

            if not sheet_has_content:
                lines.append(f"{sheet.title}: (empty sheet)")

        workbook.close()

        if not lines:
            lines.append("(no Excel content found)")

        return lines, ""

    except Exception as err:
        return None, f"Failed to read XLSX file '{file_path}': {err}"


def get_db_dump_lines(db_path):
    """Connects to a SQLite database and returns its dump as a list of strings."""
    try:
        conn, err = startup.connect_db(db_path)
        if not conn:
            return None, err

        # We strip trailing newlines to ensure clean processing by difflib
        dump_lines = [line.strip() for line in conn.iterdump()]
        conn.close()

        # Sorting is okay for database dump because order can sometimes vary.
        dump_lines.sort()

        return dump_lines, ""

    except sqlite3.Error as err:
        error_message = f"Failed to fetch data from database '{db_path}'"
        app.logger.error(
            f"error_message: '{error_message}' | database: '{db_path}' | error: {err}"
        )
        return None, error_message


def get_text_file_lines(file_path):
    """Reads normal text/code files."""
    try:
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                return [line.rstrip("\n") for line in file.readlines()], ""
        except UnicodeDecodeError:
            with open(file_path, "r", encoding="cp1252") as file:
                return [line.rstrip("\n") for line in file.readlines()], ""

    except Exception as err:
        return None, f"Failed to read text/code file '{file_path}': {err}"


def get_pdf_lines(file_path):
    """Extracts text from a PDF file."""
    if PdfReader is None:
        return (
            None,
            "PDF support requires pypdf. Install it with: python -m pip install pypdf",
        )

    try:
        reader = PdfReader(file_path)
        lines = []

        for page_index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            lines.append(f"--- Page {page_index} ---")
            lines.extend(text.splitlines())

        return lines, ""

    except Exception as err:
        return None, f"Failed to read PDF file '{file_path}': {err}"


def normalize_docx_text(text):
    """
    Normalizes DOCX text for accurate content-only comparison.

    Ignores:
    - visual line wrapping
    - tabs
    - repeated spaces
    - manual line breaks

    Keeps:
    - actual words
    - punctuation
    - citations
    - sentence content
    """
    if not text:
        return ""

    text = text.replace("\xa0", " ")
    text = text.replace("\u200b", "")
    text = text.replace("\t", " ")
    text = text.replace("\r", " ")
    text = text.replace("\n", " ")

    text = re.sub(r"\s+", " ", text)

    return text.strip()


def is_sentence_boundary(text, index):
    """
    Checks whether '.', '!' or '?' is a real sentence boundary.

    Avoids bad splits for:
    - Dr.
    - e.g.
    - i.e.
    - decimal numbers
    - citations like .[1][2][3]
    """
    char = text[index]

    if char not in ".!?":
        return False

    # Avoid decimal numbers: 3.14
    if (
        char == "."
        and index > 0
        and index + 1 < len(text)
        and text[index - 1].isdigit()
        and text[index + 1].isdigit()
    ):
        return False

    # Avoid common abbreviations
    prefix = text[: index + 1]
    abbreviation_match = re.search(r"([A-Za-z.]+)\.$", prefix)

    if abbreviation_match:
        possible_abbreviation = abbreviation_match.group(1).lower().strip(".")

        if possible_abbreviation in ABBREVIATIONS:
            return False

    j = index + 1

    # Keep citations attached to the sentence: .[1][2][3]
    while j < len(text):
        citation_match = re.match(r"\[[^\]]+\]", text[j:])

        if not citation_match:
            break

        j += citation_match.end()

    # Keep closing quotes/brackets attached
    while j < len(text) and text[j] in "\"'”’)]}":
        j += 1

    if j >= len(text):
        return True

    if not text[j].isspace():
        return False

    while j < len(text) and text[j].isspace():
        j += 1

    if j >= len(text):
        return True

    next_char = text[j]

    return next_char.isupper() or next_char.isdigit() or next_char in "\"'“‘("


def split_text_into_context_units(text):
    """
    Splits text into stable sentence-level units.

    This prevents one large paragraph from becoming fully red/green
    when only one sentence or phrase changed.
    """
    text = normalize_docx_text(text)

    if not text:
        return []

    units = []
    start = 0
    index = 0

    while index < len(text):
        if is_sentence_boundary(text, index):
            end = index + 1

            # Include citations after punctuation
            while end < len(text):
                citation_match = re.match(r"\[[^\]]+\]", text[end:])

                if not citation_match:
                    break

                end += citation_match.end()

            # Include closing quotes/brackets
            while end < len(text) and text[end] in "\"'”’)]}":
                end += 1

            unit = text[start:end].strip()

            if unit:
                units.append(unit)

            start = end

            while start < len(text) and text[start].isspace():
                start += 1

            index = start
            continue

        index += 1

    remaining = text[start:].strip()

    if remaining:
        units.append(remaining)

    return units


def collect_visible_docx_text(element):
    """
    Collects only visible text from DOCX XML.

    Ignores:
    - font
    - color
    - bold/italic/underline
    - styles
    - layout

    Keeps:
    - actual text
    - hyperlinks text
    - inserted tracked-change text
    """
    text_parts = []

    for node in element.iter():
        if node.tag == WORD_TAG + "t":
            if node.text:
                text_parts.append(node.text)

        elif node.tag == WORD_TAG + "tab":
            text_parts.append(" ")

        elif node.tag == WORD_TAG + "br":
            text_parts.append(" ")

        elif node.tag == WORD_TAG + "cr":
            text_parts.append(" ")

        elif node.tag == WORD_TAG + "noBreakHyphen":
            text_parts.append("-")

        elif node.tag == WORD_TAG + "softHyphen":
            text_parts.append("")

    return normalize_docx_text("".join(text_parts))


def extract_docx_paragraph_units(paragraph):
    """
    Converts one Word paragraph into sentence/context units.
    """
    paragraph_text = collect_visible_docx_text(paragraph)
    return split_text_into_context_units(paragraph_text)


def extract_docx_paragraph_units(paragraph):
    """
    Converts one Word paragraph into sentence/context units.
    """
    paragraph_text = collect_visible_docx_text(paragraph)
    return split_text_into_context_units(paragraph_text)


def extract_docx_part_lines(root):
    """
    Extracts paragraph text from one DOCX XML part.

    This version does not use special table extraction.
    """
    lines = []

    for paragraph in root.findall(".//w:p", DOCX_NAMESPACE):
        paragraph_units = extract_docx_paragraph_units(paragraph)
        lines.extend(paragraph_units)

    return lines


def get_docx_part_label(path):
    if path == "word/document.xml":
        return "BODY"

    if path.startswith("word/header"):
        return "HEADER"

    if path.startswith("word/footer"):
        return "FOOTER"

    if path == "word/footnotes.xml":
        return "FOOTNOTES"

    if path == "word/endnotes.xml":
        return "ENDNOTES"

    if path == "word/comments.xml":
        return "COMMENTS"

    return path


def get_docx_lines(file_path):
    """
    Professional DOCX text-content diff extractor.

    Detects real text changes while ignoring:
    - font changes
    - color changes
    - style changes
    - line wrapping
    - spacing differences
    - layout changes

    Compares:
    - body text
    - headers
    - footers
    - footnotes
    - endnotes
    - comments

    Does not use special table-row extraction.
    """
    try:
        lines = []

        text_parts_to_check = [
            "word/document.xml",
            "word/footnotes.xml",
            "word/endnotes.xml",
            "word/comments.xml",
        ]

        with zipfile.ZipFile(file_path, "r") as docx_zip:
            zip_file_names = docx_zip.namelist()

            header_footer_parts = sorted(
                file_name
                for file_name in zip_file_names
                if re.match(r"word/(header|footer)\d+\.xml$", file_name)
            )

            text_parts_to_check.extend(header_footer_parts)

            for part_path in text_parts_to_check:
                if part_path not in zip_file_names:
                    continue

                xml_content = docx_zip.read(part_path)
                root = ET.fromstring(xml_content)

                part_lines = extract_docx_part_lines(root)

                if not part_lines:
                    continue

                part_label = get_docx_part_label(part_path)

                lines.append(f"--- {part_label} ---")
                lines.extend(part_lines)

        if not lines:
            lines.append("(no text found)")

        return lines, ""

    except Exception as err:
        return None, f"Failed to read DOCX file '{file_path}': {err}"


def get_zip_listing_lines(file_path):
    """
    Compares ZIP files by listing their contents.
    It compares filenames, sizes, and CRC values.
    """
    try:
        lines = []

        with zipfile.ZipFile(file_path, "r") as zip_file:
            for item in sorted(zip_file.infolist(), key=lambda x: x.filename):
                lines.append(
                    f"{item.filename} | size={item.file_size} | crc={item.CRC}"
                )

        return lines, ""

    except Exception as err:
        return None, f"Failed to read ZIP file '{file_path}': {err}"


def get_binary_file_info_lines(file_path):
    """
    Fallback for unknown binary files.
    It cannot show line-by-line diff, but it can compare file size and checksum.
    """
    try:
        file_size = os.path.getsize(file_path)

        with open(file_path, "rb") as file:
            checksum = hashlib.md5(file.read()).hexdigest()

        mime_type, _ = mimetypes.guess_type(file_path)

        return [
            f"File: {os.path.basename(file_path)}",
            f"Size: {file_size} bytes",
            f"MD5: {checksum}",
            f"MIME type: {mime_type or 'unknown'}",
        ], ""

    except Exception as err:
        return None, f"Failed to read binary file '{file_path}': {err}"


def get_artifacts():
    conn, err = startup.connect_db(app.config["DATABASE"])
    if not conn:
        return None, err
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * from artifacts a INNER JOIN artifact_storage s ON a.id = s.artifact_id
            ORDER BY a.name, a.created_at DESC
            """)
        rows = cursor.fetchall()
        grouped_artifacts = defaultdict(list)
        grouped_commit_hashes = defaultdict(list)
        for row in rows:
            grouped_artifacts[row["name"]].append(
                {
                    "id": row["id"],
                    "commit_hash": row["commit_hash"],
                    "tags": json.loads(row["tags"]) if row["tags"] else [],
                    "checksum": row["checksum"],
                    "path": row["path"],
                    "created_at": row["created_at"],
                }
            )
            grouped_commit_hashes[row["name"]].append(row["commit_hash"])
        return grouped_artifacts, grouped_commit_hashes, ""
    except sqlite3.Error as err:
        error_message = "Failed to fetch artifacts"
        app.logger.error(f"error_message: '{error_message}' | error: {err}")
        return None, None, err
    finally:
        conn.close()


def highlight_changes(old_row, new_row):
    highlighted_old = []
    highlighted_new = []

    max_len = max(len(old_row), len(new_row))
    for i in range(max_len):
        old_val = str(old_row[i]) if i < len(old_row) else ""
        new_val = str(new_row[i]) if i < len(new_row) else ""

        if old_val != new_val:
            highlighted_old.append(
                f"<span style='background-color: #ffcccc;'>{html.escape(old_val)}</span>"
            )
            highlighted_new.append(
                f"<span style='background-color: #ccffcc;'>{html.escape(new_val)}</span>"
            )
        else:
            highlighted_old.append(html.escape(old_val))
            highlighted_new.append(html.escape(new_val))

    return f"({', '.join(highlighted_old)})", f"({', '.join(highlighted_new)})"


def get_file_diff_lines(file_path):
    ext = Path(file_path).suffix.lower()

    if ext in SQLITE_EXTENSIONS:
        return get_db_dump_lines(file_path)

    if ext in TEXT_EXTENSIONS:
        return get_text_file_lines(file_path)

    if ext in PDF_EXTENSIONS:
        return get_pdf_lines(file_path)

    if ext in DOCX_EXTENSIONS:
        return get_docx_lines(file_path)

    if ext in XLS_EXTENSIONS:
        return None, "Old .xls files are not supported. Please save the file as .xlsx."

    if ext in ZIP_EXTENSIONS:
        return get_zip_listing_lines(file_path)

    return get_binary_file_info_lines(file_path)


def calculate_diff(name, commit_hash_a, commit_hash_b):
    path_a, err = get_artifact_file_path(name, commit_hash_a)
    if err:
        app.logger.error(
            f"error_message: '{err}' | name: '{name}' | commit_hash: '{commit_hash_a}'"
        )
        return None, err

    path_b, err = get_artifact_file_path(name, commit_hash_b)
    if err:
        app.logger.error(
            f"error_message: '{err}' | name: '{name}' | commit_hash: '{commit_hash_b}'"
        )
        return None, err

    if not os.path.exists(path_a):
        error_message = (
            f"File does not exist for artifact '{name}' and commit '{commit_hash_a}'"
        )
        app.logger.error(f"error_message: '{error_message}' | path: '{path_a}'")
        return None, error_message

    if not os.path.exists(path_b):
        error_message = (
            f"File does not exist for artifact '{name}' and commit '{commit_hash_b}'"
        )
        app.logger.error(f"error_message: '{error_message}' | path: '{path_b}'")
        return None, error_message

    ext_a = Path(path_a).suffix.lower()
    ext_b = Path(path_b).suffix.lower()

    if ext_a in XLSX_EXTENSIONS and ext_b in XLSX_EXTENSIONS:
        excel_diff, err = calculate_xlsx_diff_table(path_a, path_b)

        if err:
            app.logger.error(f"error: '{err}' | path_a: '{path_a}' | path_b: '{path_b}'")
            return None, err

        return excel_diff, None

    rows_a, err = get_file_diff_lines(path_a)
    if err:
        app.logger.error(f"error: '{err}' | path: '{path_a}'")
        return None, err

    rows_b, err = get_file_diff_lines(path_b)
    if err:
        app.logger.error(f"error: '{err}' | path: '{path_b}'")
        return None, err

    if rows_a == rows_b:
        all_diffs = "(no differences found)"
    else:
        all_diffs = difflib.HtmlDiff(wrapcolumn=None).make_table(
            rows_a,
            rows_b,
            fromdesc=f"{name}@{commit_hash_a}",
            todesc=f"{name}@{commit_hash_b}",
            context=False,
        )

    return all_diffs, None


def parse_tags(tags_text):
    """
    Safely converts the tags column into a dictionary.
    If tags is empty, invalid, or not a JSON object, it returns an empty dict.
    """
    if not tags_text:
        return {}

    try:
        tags = json.loads(tags_text)
    except json.JSONDecodeError:
        return {}

    if isinstance(tags, dict):
        return tags

    # If old tags were stored as a list like ["local", "test"],
    # convert them into a dictionary so we can add deployed=True/False.
    if isinstance(tags, list):
        return {"labels": tags}

    return {}


def set_uniquely_deployed_flag(artifact_id):
    """
    Marks the selected artifact as deployed.
    Marks all other artifacts with the same name as not deployed.
    """
    conn, err = startup.connect_db(app.config["DATABASE"])
    if not conn:
        return str(err), 500

    try:
        cursor = conn.cursor()

        cursor.execute("SELECT name FROM artifacts WHERE id = ?", (artifact_id,))

        row = cursor.fetchone()

        if not row:
            return "Artifact not found", 404

        name = row["name"] if isinstance(row, sqlite3.Row) else row[0]

        cursor.execute("SELECT id, tags FROM artifacts WHERE name = ?", (name,))

        artifacts = cursor.fetchall()

        for row in artifacts:
            current_artifact_id = row["id"] if isinstance(row, sqlite3.Row) else row[0]
            tags_text = row["tags"] if isinstance(row, sqlite3.Row) else row[1]

            tags = parse_tags(tags_text)

            if current_artifact_id == artifact_id:
                tags["deployed"] = True
            else:
                tags["deployed"] = False

            cursor.execute(
                "UPDATE artifacts SET tags = ? WHERE id = ?",
                (json.dumps(tags), current_artifact_id),
            )

        conn.commit()
        return "OK", 200

    except Exception as err:
        conn.rollback()
        error_message = "Failed to update deployed flag"
        app.logger.error(
            f"error_message: '{error_message}' | artifact_id: {artifact_id} | error: {err}"
        )
        return error_message, 500

    finally:
        conn.close()


@app.route("/")
def index():
    artifacts, commit_hashes, err = get_artifacts()
    app.logger.info(f"artifacts: {commit_hashes}")
    if err:
        app.logger.error(f"error: '{err}'")
        return str(err), 500
    return render_template(
        "index.html", artifacts=artifacts, commit_hashes=commit_hashes
    )


@app.route("/instructions")
def instructions():
    return render_template("instructions.html")


@app.route("/push", methods=["GET"])
def add():
    return render_template("add.html")


@app.route("/push", methods=["POST"])
def push():
    provided_key = request.headers.get("X-API-Key", "")
    expected_key = app.config.get("DEPLOY_KEY", "")

    if not expected_key:
        app.logger.error("DEPLOY_KEY is not configured")
        return "Server API key is not configured", 500

    if not provided_key:
        app.logger.warning("Push rejected: API key is missing")
        return "API key is missing", 401

    if not hmac.compare_digest(provided_key, expected_key):
        app.logger.warning("Push rejected: invalid API key")
        return "API key is invalid", 401

    name = request.form.get("name")
    commit_hash = request.form.get("commit_hash")
    tags = request.form.get("tags", "")
    file_content = request.files.get("file")

    if not name or not commit_hash or not file_content:
        error_message = "Name and commit hash and file are required"
        app.logger.error(
            f"error: '{error_message}' | name: '{name}' | "
            f"commit_hash: '{commit_hash}' | file_content: '{file_content}'"
        )
        return error_message, 400

    is_allowed_characters = re.match(r"^[a-zA-Z0-9_.-]+$", name)
    if not is_allowed_characters:
        error_message = "Invalid name. Name must contain only: a-z, A-Z, 0-9, dot, dash, underscore."
        app.logger.error(f"error_message: '{error_message}' | name: '{name}'")
        return error_message, 400

    try:
        parsed_tags = json.loads(tags) if tags else {}
    except json.JSONDecodeError as err:
        error_message = "Invalid tags: Must be a valid JSON-encoded string."
        app.logger.error(
            f"error_message: '{error_message}' | tags: '{tags}' | error: {err}"
        )
        return error_message, 400

    conn, err = startup.connect_db(app.config["DATABASE"])
    if not conn:
        return str(err), 500

    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO artifacts (name, commit_hash, tags) VALUES (?, ?, ?)
            """,
            (name, commit_hash, tags),
        )

        artifact_id = cursor.lastrowid

        _, ext = os.path.splitext(file_content.filename)

        filepath = os.path.join(
            app.config["STORAGE_FOLDER"], f"{name}_{commit_hash}{ext}"
        )

        file_content.save(filepath)

        with open(filepath, "rb") as f:
            checksum = hashlib.md5(f.read()).hexdigest()

        cursor.execute(
            """
            INSERT INTO artifact_storage (artifact_id, checksum, path) VALUES (?, ?, ?)
            """,
            (artifact_id, checksum, filepath),
        )

        conn.commit()
        cursor.close()

    except Exception as err:
        conn.rollback()
        error_message = "Failed to push artifact"
        app.logger.error(f"error_message: '{error_message}' | error: {err}")
        return error_message, 500

    finally:
        conn.close()

    return "OK", 201


@app.route("/download", methods=["GET"])
def download_api():
    name = request.args.get("name")
    commit_hash = request.args.get("commit_hash", "")

    if not name:
        error_message = "Name is required"
        app.logger.error(f"error: '{error_message}' | name: '{name}'")
        return error_message, 400

    conn, err = startup.connect_db(app.config["DATABASE"])
    if not conn:
        return str(err), 500

    try:
        cursor = conn.cursor()

        if commit_hash:
            cursor.execute(
                """
                SELECT path FROM artifact_storage
                JOIN artifacts ON artifact_storage.artifact_id = artifacts.id
                WHERE artifacts.name = ? AND artifacts.commit_hash = ?
                """,
                (name, commit_hash),
            )
        else:
            # if commit hash is not specified, we download the latest artifact.
            cursor.execute(
                """
                SELECT path FROM artifact_storage
                JOIN artifacts ON artifact_storage.artifact_id = artifacts.id
                WHERE artifacts.name = ? order by artifacts.id desc
                """,
                (name,),
            )

        row = cursor.fetchone()
        if not row:
            error_message = "Artifact not found"
            app.logger.error(
                f"error_message: '{error_message}' | name: '{name}' | commit_hash: '{commit_hash}'"
            )
            return error_message, 404
        filepath = row["path"]

        return send_file(filepath, as_attachment=True), 200
    except Exception as err:
        error_message = "Failed to download artifact"
        app.logger.error(f"error_message: '{error_message}' | error: {err}")
        return error_message, 500
    finally:
        conn.close()


@app.route("/download/<int:artifact_id>")
def download_dashboard(artifact_id):
    conn, err = startup.connect_db(app.config["DATABASE"])
    if not conn:
        return str(err), 500

    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT path FROM artifact_storage WHERE artifact_id = ?
            """,
            (artifact_id,),
        )
        row = cursor.fetchone()
        if not row:
            error_message = "Artifact not found"
            app.logger.error(
                f"error_message: '{error_message}' | artifact_id: {artifact_id}"
            )
            return error_message, 404
        filepath = row["path"]

        return send_file(filepath, as_attachment=True), 200
    except Exception as err:
        error_message = "Failed to download artifact"
        app.logger.error(f"error_message: '{error_message}' | error: {err}")
        return error_message, 500
    finally:
        conn.close()


@app.route("/diff")
def diff_api():
    name = request.args.get("name")
    commit_hash_a = request.args.get("commit_hash_a")
    commit_hash_b = request.args.get("commit_hash_b")

    if not name or not commit_hash_a or not commit_hash_b:
        error_message = "Name, both commit hashes and table are required"
        app.logger.error(
            f"error: '{error_message}' | name: '{name}' | commit_hash_a: '{commit_hash_a}' | commit_hash_b: '{commit_hash_b}'"
        )
        return error_message, 400

    diff, err = calculate_diff(name, commit_hash_a, commit_hash_b)

    if err:
        app.logger.error(
            f"error: '{err}' | name: '{name}' | commit_hash_a: '{commit_hash_a}' | commit_hash_b: '{commit_hash_b}'"
        )
        return str(err), 500

    return jsonify(
        {
            "name1": name,
            "commit_hash_a": commit_hash_a,
            "commit_hash_b": commit_hash_b,
            "diff_by_table": diff,
        }
    )


@app.route("/deploy/<int:artifact_id>")
def deploy_api(artifact_id):
    """
    Marks this artifact as deployed.
    All other artifacts with the same name become undeployed.
    """
    return set_uniquely_deployed_flag(artifact_id)


if __name__ == "__main__":
    app.run(debug=True)
